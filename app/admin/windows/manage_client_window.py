import os
from PyQt6.QtWidgets import (QApplication, QMessageBox, QFileDialog,
                            QTableWidgetItem)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as XLImage

from app.admin.interfaces.manage_client_ui import ManageClient
from app.admin.interfaces.add_client_ui import AddClientWindow
from app.admin.interfaces.memb_edit_ui import MembEdit
from app.admin.interfaces.addFreeze_ui import AddFreeze
from app.admin.db_objects.database_for_me import Database

db = Database("localhost", "root", "", "fitness")


class ManageClientWindow(ManageClient):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.current_client_id = None

        self._setup_connections()
        self.load_clients()

    def _setup_connections(self):
        self.btnAdd.clicked.connect(self.add_client)
        self.btnEdit.clicked.connect(self.edit_membership)
        self.btnCard.clicked.connect(self.print_club_card)
        self.btnFreeze.clicked.connect(self.freeze_membership)
        self.btnBack.clicked.connect(self.go_back)
        
        self.tableWidget.itemSelectionChanged.connect(self.on_client_selected)
        
        self.btnEdit.setEnabled(False)
        self.btnCard.setEnabled(False)
        self.btnFreeze.setEnabled(False)

    def load_clients(self):
        """Load all clients into the table"""
        try:
            with db.conn.cursor() as cursor:
                cursor.callproc('GetAllClients')
                
                # Clear the table
                self.tableWidget.setRowCount(0)
                
                # Get the first result set from the stored procedure
                result = cursor.fetchall()
                
                # If no results, try to get them from stored_results()
                if not result:
                    for result_set in cursor.stored_results():
                        result = result_set.fetchall()
                        if result:
                            break
                
                # Populate the table
                for row in result:
                    row_position = self.tableWidget.rowCount()
                    self.tableWidget.insertRow(row_position)
                    
                    # Add client data to the table
                    self.tableWidget.setItem(row_position, 0, QTableWidgetItem(str(row.get('userID', ''))))
                    self.tableWidget.setItem(row_position, 1, QTableWidgetItem(row.get('last_name', '')))
                    self.tableWidget.setItem(row_position, 2, QTableWidgetItem(row.get('first_name', '')))
                    self.tableWidget.setItem(row_position, 3, QTableWidgetItem(row.get('phone', '')))
                    
                    # Format membership info
                    active_memberships = row.get('active_memberships', 0)
                    latest_end = row.get('latest_membership_end', '')
                    membership_info = f"{active_memberships} активных"
                    if latest_end:
                        membership_info += f"\nДо {latest_end.strftime('%d.%m.%Y') if hasattr(latest_end, 'strftime') else latest_end}"
                    
                    self.tableWidget.setItem(row_position, 4, QTableWidgetItem(membership_info))
                
                # Adjust column widths
                self.tableWidget.resizeColumnsToContents()
                
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось загрузить список клиентов: {str(e)}')

    def on_client_selected(self):
        """Handle client selection changes"""
        selected_items = self.tableWidget.selectedItems()
        if selected_items and selected_items[0]:
            row = selected_items[0].row()
            self.current_client_id = int(self.tableWidget.item(row, 0).text())
            self.btnEdit.setEnabled(True)
            self.btnCard.setEnabled(True)
            self.btnFreeze.setEnabled(True)
        else:
            self.current_client_id = None
            self.btnEdit.setEnabled(False)
            self.btnCard.setEnabled(False)
            self.btnFreeze.setEnabled(False)

    def add_client(self):
        """Open the add client dialog"""
        self.add_client_window = AddClientWindow(self)
        self.add_client_window.client_added.connect(self._create_client)
        self.add_client_window.back_requested.connect(self._return_from_child)
        self.hide()
        self.add_client_window.show()

    def _create_client(self, data):
        try:
            with db.conn.cursor() as cursor:
                first_name = data.get('first_name', '')
                last_name = data.get('last_name', '')
                middle_name = data.get('middle_name', '')
                phone = data.get('phone', '')
                email = data.get('email', '')
                birth_date = data.get('birth_date')
                health_limits = data.get('health_limits', '')

                login = phone or f"{first_name.lower()}_{last_name.lower()}"
                password = phone or '123456'

                cursor.callproc('AddClient', (
                    first_name,
                    last_name,
                    middle_name,
                    phone,
                    email,
                    birth_date,
                    health_limits,
                    login,
                    password
                ))
            db.conn.commit()
            if hasattr(self.add_client_window, 'save_succeeded'):
                self.add_client_window.save_succeeded = True
            QMessageBox.information(self, 'Успех', 'Клиент успешно добавлен')
            self.load_clients()
        except Exception as e:
            db.conn.rollback()
            QMessageBox.critical(self, 'Ошибка', f'Не удалось добавить клиента: {str(e)}')

    def edit_membership(self):
        """Open the membership edit dialog for the selected client"""
        if not self.current_client_id:
            QMessageBox.warning(self, 'Ошибка', 'Пожалуйста, выберите клиента')
            return
            
        self.membership_edit_window = MembEdit(self.current_client_id, self)
        self.membership_edit_window.membership_updated.connect(self.load_clients)
        self.membership_edit_window.back_requested.connect(self._return_from_child)
        self.hide()
        self.membership_edit_window.show()

    def print_club_card(self):
        """Generate an Excel file with the client's club card"""
        if not self.current_client_id:
            return
            
        try:
            with db.conn.cursor() as cursor:
                # Get client info
                cursor.execute("""
                    SELECT userID, first_name, last_name, middle_name, phone, email,
                           birthDate, health_limits
                    FROM Users 
                    WHERE userID = %s
                """, (self.current_client_id,))
                
                client = cursor.fetchone()
                if not client:
                    QMessageBox.warning(self, 'Ошибка', 'Клиент не найден')
                    return
                
                # Get active memberships
                cursor.callproc('GetClientMemberships', (self.current_client_id,))
                memberships = cursor.fetchall()
                
                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Клубная карта"
                
                # Add logo (if exists)
                logo_path = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'static', 'logo.png')
                if os.path.exists(logo_path):
                    try:
                        img = XLImage(logo_path)
                        img.width = 100
                        img.height = 50
                        ws.add_image(img, 'A1')
                    except Exception as e:
                        print(f"Could not add logo: {e}")
                
                # Add title
                ws['A3'] = "КЛУБНАЯ КАРТА"
                ws['A3'].font = Font(size=16, bold=True)
                ws.merge_cells('A3:D3')
                
                # Add client info
                ws['A5'] = "ФИО:"
                ws['B5'] = f"{client.get('last_name', '')} {client.get('first_name', '')} {client.get('middle_name', '')}"
                ws.merge_cells('B5:D5')
                
                ws['A6'] = "Телефон:"
                ws['B6'] = client.get('phone', '')
                ws.merge_cells('B6:D6')
                
                ws['A7'] = "Дата рождения:"
                birth_date = client.get('birthDate')
                ws['B7'] = birth_date.strftime('%d.%m.%Y') if hasattr(birth_date, 'strftime') else str(birth_date or '')
                ws.merge_cells('B7:D7')
                
                # Add memberships
                if memberships:
                    ws['A9'] = "Абонементы:"
                    ws['A9'].font = Font(bold=True)
                    
                    headers = ['Тип', 'Начало', 'Окончание', 'Статус', 'Осталось посещений']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=10, column=col)
                        cell.value = header
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                    
                    for i, memb in enumerate(memberships, 11):
                        ws.cell(row=i, column=1, value=memb.get('membType', ''))
                        
                        start_date = memb.get('startDate')
                        end_date = memb.get('endDate')
                        
                        ws.cell(row=i, column=2, value=start_date.strftime('%d.%m.%Y') 
                              if hasattr(start_date, 'strftime') else str(start_date or ''))
                        
                        ws.cell(row=i, column=3, value=end_date.strftime('%d.%m.%Y') 
                              if hasattr(end_date, 'strftime') else str(end_date or ''))
                        
                        ws.cell(row=i, column=4, value=memb.get('status_text', ''))
                        
                        visits_used = memb.get('visitsUsed', 0)
                        visits_total = memb.get('visitsTotal', 0)
                        ws.cell(row=i, column=5, value=f"{visits_total - visits_used} из {visits_total}")
                else:
                    ws['A9'] = "Нет активных абонементов"
                
                # Adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value or '')) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min((max_length + 2) * 1.2, 30)  # Limit max width
                    ws.column_dimensions[column].width = adjusted_width
                
                # Save the file
                file_name = f"Клубная карта {client.get('last_name', '')} {client.get('first_name', '')}.xlsx"
                file_path, _ = QFileDialog.getSaveFileName(
                    self, 
                    'Сохранить клубную карту', 
                    os.path.expanduser(f'~/{file_name}'),
                    'Excel Files (*.xlsx)'
                )
                
                if file_path:
                    if not file_path.endswith('.xlsx'):
                        file_path += '.xlsx'
                    wb.save(file_path)
                    QMessageBox.information(self, 'Успех', f'Клубная карта сохранена: {file_path}')
                
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось сгенерировать клубную карту: {str(e)}')

    def freeze_membership(self):
        """Open the freeze membership dialog"""
        if not self.current_client_id:
            QMessageBox.warning(self, 'Ошибка', 'Пожалуйста, выберите клиента')
            return
                
        self.freeze_window = AddFreeze(self.current_client_id, self)
        self.freeze_window.freeze_added.connect(self.load_clients)
        self.freeze_window.back_requested.connect(self._return_from_child)
        self.hide()
        self.freeze_window.show()

    def go_back(self):
        if self.parent_window:
            self.parent_window.show()
        self.close()

    def closeEvent(self, event):
        if self.parent_window:
            self.parent_window.show()
        event.accept()

    def _return_from_child(self):
        self.load_clients()
        self.show()

if __name__ == "__main__":
    app = QApplication([])
    window = ManageClientWindow()
    window.show()
    app.exec()
